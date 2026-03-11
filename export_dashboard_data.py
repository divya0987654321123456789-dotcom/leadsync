from __future__ import annotations

import json
from collections import Counter, defaultdict
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parent
SOURCE_XLSX = ROOT / "2025-2026 Lead Generation Data.xlsx"
OUTPUT_DIR = ROOT / "data"
OUTPUT_FILE = OUTPUT_DIR / "dashboard-data.js"
MONTH_ORDER = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def load_sheet_rows(workbook: openpyxl.Workbook, sheet_name: str) -> list[dict[str, Any]]:
    worksheet = workbook[sheet_name]
    headers = [clean_value(cell) for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    valid_columns = [(index, header) for index, header in enumerate(headers) if header]

    rows: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        item = {}
        for index, header in valid_columns:
            item[header] = clean_value(row[index] if index < len(row) else None)
        if any(value is not None for value in item.values()):
            rows.append(item)
    return rows


def counter_to_pairs(counter: Counter, limit: int | None = None) -> list[dict[str, Any]]:
    items = counter.most_common(limit)
    return [{"label": key, "value": value} for key, value in items]


def sort_month_counts(month_counts: dict[str, int]) -> list[dict[str, Any]]:
    return [{"month": month, "value": month_counts.get(month, 0)} for month in MONTH_ORDER if month in month_counts]


def safe_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def build_leads_summary(rows: list[dict[str, Any]], sheet_name: str) -> dict[str, Any]:
    by_month: dict[str, int] = defaultdict(int)
    by_source: Counter = Counter()
    by_stage: Counter = Counter()
    by_type: Counter = Counter()
    by_response: Counter = Counter()
    by_state: Counter = Counter()
    by_owner: Counter = Counter()
    by_campaign: Counter = Counter()
    by_campaign_type: Counter = Counter()
    revenue_total = 0.0
    positive_count = 0
    negative_count = 0

    spotlight_rows = []
    for row in rows:
        by_month[row.get("Month") or "Unknown"] += 1
        by_source[row.get("Lead Source") or "Unknown"] += 1
        by_stage[row.get("Lead Stage") or "Unknown"] += 1
        by_type[row.get("Lead type") or "Unknown"] += 1
        by_response[row.get("Response Type") or "Unknown"] += 1
        by_state[row.get("State") or "Unknown"] += 1
        by_owner[row.get("Assign to") or "Unassigned"] += 1
        by_campaign[row.get("Campaign name") or "Unknown"] += 1
        by_campaign_type[row.get("Campaign Type") or "Unknown"] += 1
        revenue_total += safe_number(row.get("Revenue"))

        response = (row.get("Response Type") or "").lower()
        if response == "positive":
            positive_count += 1
        elif response == "negative":
            negative_count += 1

        spotlight_rows.append(
            {
                "date": row.get("Date"),
                "account": row.get("Account Name") or "Unknown",
                "campaign": row.get("Campaign name") or "Unknown",
                "stage": row.get("Lead Stage") or "Unknown",
                "responseType": row.get("Response Type") or "Unknown",
                "owner": row.get("Assign to") or "Unassigned",
                "state": row.get("State") or "Unknown",
                "outcome": row.get("Outcome"),
            }
        )

    total = len(rows)
    return {
        "sheetName": sheet_name,
        "totalRows": total,
        "revenueTotal": revenue_total,
        "positiveRate": round((positive_count / total) * 100, 1) if total else 0,
        "negativeRate": round((negative_count / total) * 100, 1) if total else 0,
        "byMonth": sort_month_counts(by_month),
        "bySource": counter_to_pairs(by_source),
        "byStage": counter_to_pairs(by_stage),
        "byLeadType": counter_to_pairs(by_type),
        "byResponseType": counter_to_pairs(by_response),
        "byState": counter_to_pairs(by_state, 12),
        "byOwner": counter_to_pairs(by_owner, 12),
        "byCampaign": counter_to_pairs(by_campaign, 10),
        "byCampaignType": counter_to_pairs(by_campaign_type, 10),
        "recentLeads": sorted(spotlight_rows, key=lambda item: item["date"] or "", reverse=True)[:12],
    }


def build_email_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    by_month: dict[str, int] = defaultdict(int)
    by_stage: Counter = Counter()
    by_type: Counter = Counter()
    by_state: Counter = Counter()
    by_owner: Counter = Counter()
    by_campaign: Counter = Counter()
    by_campaign_type: Counter = Counter()

    campaign_stage_rollup: dict[str, Counter] = defaultdict(Counter)
    for row in rows:
        month = row.get("Month") or "Unknown"
        stage = row.get("Contact Stage") or "Unknown"
        campaign = row.get("Campaign name") or "Unknown"
        campaign_type = row.get("Campaign Type") or "Unknown"

        by_month[month] += 1
        by_stage[stage] += 1
        by_type[row.get("Lead type") or "Unknown"] += 1
        by_state[row.get("State") or "Unknown"] += 1
        by_owner[row.get("Sales Owner") or "Unassigned"] += 1
        by_campaign[campaign] += 1
        by_campaign_type[campaign_type] += 1
        campaign_stage_rollup[campaign][stage] += 1

    total = len(rows)
    delivered = by_stage.get("Delivered", 0)
    opened = by_stage.get("Opened", 0)
    replied = by_stage.get("Replied", 0)
    bounced = by_stage.get("Bounced", 0)

    top_campaigns = []
    for campaign, count in by_campaign.most_common(12):
        top_campaigns.append(
            {
                "label": campaign,
                "value": count,
                "stages": counter_to_pairs(campaign_stage_rollup[campaign]),
            }
        )

    return {
        "totalRows": total,
        "deliveredRate": round((delivered / total) * 100, 1) if total else 0,
        "openRate": round((opened / total) * 100, 1) if total else 0,
        "replyRate": round((replied / total) * 100, 1) if total else 0,
        "bounceRate": round((bounced / total) * 100, 1) if total else 0,
        "byMonth": sort_month_counts(by_month),
        "byStage": counter_to_pairs(by_stage),
        "byLeadType": counter_to_pairs(by_type),
        "byState": counter_to_pairs(by_state, 12),
        "byOwner": counter_to_pairs(by_owner, 12),
        "byCampaign": top_campaigns,
        "byCampaignType": counter_to_pairs(by_campaign_type, 10),
    }


def build_global_summary(lead_summaries: list[dict[str, Any]], email_summary: dict[str, Any]) -> dict[str, Any]:
    total_leads = sum(item["totalRows"] for item in lead_summaries)
    positive_leads = sum(round(item["totalRows"] * item["positiveRate"] / 100) for item in lead_summaries)
    return {
        "totalLeadRows": total_leads,
        "positiveLeadRows": positive_leads,
        "emailRows": email_summary["totalRows"],
        "topLineConversion": round((positive_leads / total_leads) * 100, 1) if total_leads else 0,
    }


def main() -> None:
    workbook = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    leads_2026 = load_sheet_rows(workbook, "2026 Leads")
    leads_2025 = load_sheet_rows(workbook, "2025 Leads")
    email_campaign = load_sheet_rows(workbook, "Email Campaign 2025")

    summary_2026 = build_leads_summary(leads_2026, "2026 Leads")
    summary_2025 = build_leads_summary(leads_2025, "2025 Leads")
    email_summary = build_email_summary(email_campaign)

    payload = {
        "generatedAt": datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "globalSummary": build_global_summary([summary_2026, summary_2025], email_summary),
        "leadSheets": {
            "2026 Leads": summary_2026,
            "2025 Leads": summary_2025,
        },
        "emailCampaign": email_summary,
    }

    OUTPUT_DIR.mkdir(exist_ok=True)
    output = "window.__DASHBOARD_DATA__ = " + json.dumps(payload, separators=(",", ":")) + ";"
    OUTPUT_FILE.write_text(output, encoding="utf-8")


if __name__ == "__main__":
    main()
