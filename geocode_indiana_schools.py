from __future__ import annotations

import csv
import json
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
INPUT_CSV = BASE_DIR / 'indiana_schools_maps_links.csv'
INPUT_XLSX = BASE_DIR / 'Indiana Schools Coordinates.xlsx'
OUTPUT_CSV = BASE_DIR / 'indiana_schools_maps_links_with_coordinates.csv'
USER_AGENT = 'codex-cli/1.0 (school geocoding task)'


def load_rows() -> list[dict[str, str]]:
    if INPUT_CSV.exists():
        with INPUT_CSV.open('r', encoding='utf-8-sig', newline='') as infile:
            return list(csv.DictReader(infile))

    if INPUT_XLSX.exists():
        workbook = load_workbook(INPUT_XLSX, read_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        headers = [str(cell) if cell is not None else '' for cell in rows[0]]
        records: list[dict[str, str]] = []
        for raw_row in rows[1:]:
            record: dict[str, str] = {}
            for index, header in enumerate(headers):
                value = raw_row[index] if index < len(raw_row) else ''
                record[header] = '' if value is None else str(value)
            records.append(record)
        return records

    raise FileNotFoundError(f'No input file found at {INPUT_CSV} or {INPUT_XLSX}')


def extract_query_address(row: dict[str, str]) -> str:
    google_maps_link = row.get('Google_Maps_Link', '').strip()
    if google_maps_link:
        parsed = urllib.parse.urlparse(google_maps_link)
        query_values = urllib.parse.parse_qs(parsed.query)
        query_address = query_values.get('query', [''])[0].strip()
        if query_address:
            return query_address

    return row.get('Full_Address', '').strip()


def geocode(address: str) -> tuple[str, str, str]:
    if not address:
        return '', '', 'missing_address'

    params = {
        'q': address,
        'format': 'jsonv2',
        'limit': 1,
        'addressdetails': 0,
    }
    url = 'https://nominatim.openstreetmap.org/search?' + urllib.parse.urlencode(params)
    request = urllib.request.Request(url, headers={'User-Agent': USER_AGENT})

    with urllib.request.urlopen(request, timeout=30) as response:
        payload = json.load(response)

    if not payload:
        return '', '', 'not_found'

    best = payload[0]
    return str(best.get('lat', '')), str(best.get('lon', '')), 'ok'


def main() -> int:
    try:
        rows = load_rows()
    except FileNotFoundError as exc:
        print(str(exc), file=sys.stderr)
        return 1

    if not rows:
        print('No rows found in input file.', file=sys.stderr)
        return 1

    output_rows: list[dict[str, str]] = []
    for index, row in enumerate(rows, start=1):
        search_address = extract_query_address(row)
        try:
            latitude, longitude, status = geocode(search_address)
        except Exception as exc:
            latitude, longitude, status = '', '', f'error: {exc}'

        output_row = dict(row)
        output_row['Search_Address'] = search_address
        output_row['Latitude'] = latitude
        output_row['Longitude'] = longitude
        output_row['Geocode_Status'] = status
        output_rows.append(output_row)

        company = row.get('Company', '').strip()
        print(f'[{index}/{len(rows)}] {company} -> {status}')
        time.sleep(1.0)

    fieldnames = list(output_rows[0].keys())
    with OUTPUT_CSV.open('w', encoding='utf-8', newline='') as outfile:
        writer = csv.DictWriter(outfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(output_rows)

    print(f'Saved output to {OUTPUT_CSV}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
